#!/usr/bin/env python3
"""Quick script to read Signatera Excel file"""
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(["pip3", "install", "openpyxl"])
    import openpyxl

wb = openpyxl.load_workbook('/Users/JDKristenson/Desktop/JM_Signatera_variant_information_19053697.xlsx')
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"\nTotal rows: {ws.max_row}, Total cols: {ws.max_column}")

print("\n=== Column Headers ===")
headers = [cell.value for cell in ws[1]]
for i, header in enumerate(headers, 1):
    print(f"Column {chr(64+i)}: {header}")

print("\n=== All 16 Rows (markers) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=17, values_only=True), 2):
    print(f"\nRow {row_idx}:")
    for col_idx, value in enumerate(row):
        header = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx+1}"
        print(f"  {header}: {value}")

